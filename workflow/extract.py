from openpyxl import load_workbook


def extract_data():
    path = "data/online_retail.xlsx"

    workbook = load_workbook(path, read_only=True, rich_text=True)
    sheet = workbook.active

    return sheet.iter_rows(min_row=2, values_only=True)
